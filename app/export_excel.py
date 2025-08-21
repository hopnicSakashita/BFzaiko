import os
from dotenv import load_dotenv
import openpyxl
from openpyxl.utils import get_column_letter
import shutil
from datetime import datetime
import traceback

from app.models import log_error
from app.constants import ExcelConstants, ErrorMessages

def write_to_proc_excel(shipment_date, data, output_dir=None):
    """コーティング工程管理表にデータを書き込む
    
    Args:
        shipment_date: 出荷日
        data: 書き込むデータ
        output_dir: 出力先ディレクトリ（指定がない場合は日付フォルダに保存）
    """
    try:
        # .envファイルの読み込み
        load_dotenv()
        template_path = os.getenv('PROC_EXCEL_PATH')
        
        if not template_path or not os.path.exists(template_path):
            raise FileNotFoundError(ErrorMessages.FILE_NOT_FOUND + template_path)

        # テンプレートファイルの検証
        try:
            # 一時的にファイルを開いて検証
            wb_test = openpyxl.load_workbook(template_path)
            if ExcelConstants.TEMPLATE_SHEET_NAME not in wb_test.sheetnames:
                raise ValueError(ErrorMessages.FILE_INVALID_FORMAT)
            wb_test.close()
        except Exception as e:
            raise ValueError(ErrorMessages.EXCEL_TEMPLATE_INVALID)

        # 出力先ディレクトリの設定
        if output_dir is None:
            date_str = shipment_date.strftime('%Y%m%d')
            output_dir = os.path.join(os.path.dirname(template_path), date_str)
        
        # 出力先ディレクトリの作成
        os.makedirs(output_dir, exist_ok=True)
        
        # 新しいファイル名の生成
        date_str = shipment_date.strftime('%Y%m%d')
        output_path = os.path.join(output_dir, ExcelConstants.HARDCOAT_FILENAME_TEMPLATE.format(date_str))
        
        # テンプレートファイルをコピー
        shutil.copy2(template_path, output_path)
        
        try:
            # Excelファイルを開く
            wb = openpyxl.load_workbook(output_path)
            template_sheet = wb[ExcelConstants.TEMPLATE_SHEET_NAME]
            
            # テンプレートシートを保持
            template_sheet = wb[ExcelConstants.TEMPLATE_SHEET_NAME]
            # 作業用シートを作成
            current_sheet = wb.copy_worksheet(template_sheet)
            current_sheet.title = '1'
            # カウンターの初期化
            kago = 0
            sheet_count = 1
            row_count = ExcelConstants.WRITE_START_ROW  # 書き込み開始行
            
            for item in data:
                kago += 1
                # 最大行数を超える場合、新しいシートを作成
                if row_count > ExcelConstants.WRITE_START_ROW + ExcelConstants.MAX_ROWS_PER_SHEET - 1:
                    sheet_count += 1
                    new_sheet_name = f'{sheet_count}'
                    current_sheet = wb.copy_worksheet(template_sheet)
                    current_sheet.title = new_sheet_name
                    row_count = ExcelConstants.WRITE_START_ROW  # 行カウントをリセット
                
                # データの書き込み
                current_sheet[f'{ExcelConstants.COLUMN_BASKET_NO}{row_count}'] = kago
                current_sheet[f'{ExcelConstants.COLUMN_LOT}{row_count}'] = item.lot
                current_sheet[f'{ExcelConstants.COLUMN_BASE}{row_count}'] = float(item.base) if item.base else None
                current_sheet[f'{ExcelConstants.COLUMN_ADP}{row_count}'] = float(item.adp) if item.adp else None
                current_sheet[f'{ExcelConstants.COLUMN_LR}{row_count}'] = item.lr
                current_sheet[f'{ExcelConstants.COLUMN_COLOR}{row_count}'] = item.color
                current_sheet[f'{ExcelConstants.COLUMN_QUANTITY}{row_count}'] = int(item.quantity) if item.quantity else None
                current_sheet[f'{ExcelConstants.COLUMN_BSHK_ID}{row_count}'] = item.bshk_id
                
                row_count += 1
                
            # 原紙シートを削除
            if ExcelConstants.TEMPLATE_SHEET_NAME in wb.sheetnames:
                wb.remove(wb[ExcelConstants.TEMPLATE_SHEET_NAME])
            
            # ファイルを保存して閉じる
            wb.save(output_path)
            wb.close()
            
            return output_path
            
        except Exception as e:
            # ファイル操作中のエラーの場合、作成したファイルを削除
            if os.path.exists(output_path):
                try:
                    os.remove(output_path)
                except:
                    pass
            error_msg = str(e)
            if "Permission denied" in error_msg:
                raise Exception(ErrorMessages.FILE_PERMISSION_DENIED)
            elif "Memory" in error_msg:
                raise Exception(ErrorMessages.EXCEL_MEMORY_ERROR)
            else:
                raise Exception(f"{ErrorMessages.EXCEL_CREATION_ERROR}: {error_msg}")
            
    except Exception as e:
        error_msg = str(e)
        log_error(f"コーティング工程管理表の作成中にエラーが発生しました: {error_msg}\n{traceback.format_exc()}")
        
        # エラーメッセージをより具体的に
        if isinstance(e, FileNotFoundError):
            error_msg = ErrorMessages.FILE_NOT_FOUND
        elif "Permission denied" in error_msg:
            error_msg = ErrorMessages.FILE_PERMISSION_DENIED
        elif "Memory" in error_msg:
            error_msg = ErrorMessages.EXCEL_MEMORY_ERROR
        elif "invalid literal for int()" in error_msg:
            error_msg = ErrorMessages.EXCEL_NUMBER_CONVERSION_ERROR
            
        raise Exception(f'{ErrorMessages.FILE_CREATION_FAILED}: {error_msg}')
