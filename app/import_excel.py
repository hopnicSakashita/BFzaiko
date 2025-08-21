
import tempfile
import os
import pandas as pd
import openpyxl
from datetime import datetime
import traceback

from flask import flash, redirect, url_for
from sqlalchemy import text

from app.constants import DatabaseConstants, ExcelConstants
from app.database import get_db_session
        
def import_excel_hardcoat(file):
    temp_dir = tempfile.gettempdir()
    temp_path = os.path.join(temp_dir, file.filename)
    file.save(temp_path)
    
    # ファイルサイズをチェック
    if os.path.getsize(temp_path) == 0:
        os.remove(temp_path)
        flash('ファイルが空です', 'error')
        return redirect(url_for('hardcoat_read_excel'))
        
    # Excelファイルの処理
    
    # Excelファイルを読み込む
    wb = openpyxl.load_workbook(temp_path)
    
    # データを処理
    session = get_db_session()
    processed_count = 0
    error_count = 0
    error_messages = []
    
    try:
        # 製品マスタデータを一括取得
        product_master = {}
        sql = text("""
            SELECT BFSP_PRD_ID, BFSP_BASE, BFSP_ADP, BFSP_LR, BFSP_CLR
            FROM BFSP_MST
        """)
        for row in session.execute(sql):
            key = (row.BFSP_BASE, row.BFSP_ADP, row.BFSP_LR, row.BFSP_CLR)
            product_master[key] = row.BFSP_PRD_ID

        # 既存のロットデータを一括取得
        existing_lots = set()
        sql = text("""
            SELECT DISTINCT BPDD_PRD_ID, BPDD_LOT
            FROM BPRD_DAT
            WHERE BPDD_PROC = :proc_hardcoat AND BPDD_FLG = :flg_not_shipped
        """)
        for row in session.execute(sql, {'proc_hardcoat': DatabaseConstants.PROC_HARD_COAT, 'flg_not_shipped': DatabaseConstants.BPDD_FLG_NOT_SHIPPED}):
            existing_lots.add((row.BPDD_PRD_ID, row.BPDD_LOT))

        # 各シートを処理
        for sheet_name in wb.sheetnames:
            if sheet_name == '原紙':
                continue
                
            sheet = wb[sheet_name]
            data_to_insert = []
            
            # データ行を一括で取得
            data_rows = []
            for row in range(5, sheet.max_row + 1):
                if not sheet[f'A{row}'].value:
                    break
                data_rows.append(row)

            # データを一括で処理
            for row in data_rows:
                try:
                    lot = str(sheet[f'{ExcelConstants.COLUMN_LOT}{row}'].value)  # 数値を文字列に変換
                    base = int(sheet[f'{ExcelConstants.COLUMN_BASE}{row}'].value)
                    adp = int(sheet[f'{ExcelConstants.COLUMN_ADP}{row}'].value)
                    lr = sheet[f'{ExcelConstants.COLUMN_LR}{row}'].value
                    color = sheet[f'{ExcelConstants.COLUMN_COLOR}{row}'].value
                    quantity = int(sheet[f'{ExcelConstants.COLUMN_PASS_QTY}{row}'].value)
                    bshk_id = int(sheet[f'{ExcelConstants.COLUMN_BSHK_ID}{row}'].value) if sheet[f'{ExcelConstants.COLUMN_BSHK_ID}{row}'].value else 0
                    
                    if quantity is None or quantity == 0:
                        continue

                    coat_date = sheet[f'{ExcelConstants.COLUMN_COAT_DATE}{row}'].value
                    if coat_date is None:
                        error_msg = f'シート「{sheet_name}」行{row}: コート日が入力されていません: ロット={lot}, ベース={base}, 加入度数={adp}, L/R={lr}, 色={color}'
                        error_messages.append(error_msg)
                        error_count += 1
                        continue
                        
                    if isinstance(coat_date, str):
                        try:
                            coat_dt = datetime.strptime(coat_date, '%Y/%m/%d')
                        except ValueError:
                            error_msg = f'シート「{sheet_name}」行{row}: コート日の形式が正しくありません: {coat_date} (期待形式: YYYY/MM/DD)'
                            error_messages.append(error_msg)
                            error_count += 1
                            continue
                    else:
                        coat_dt = coat_date
                    coat_dt_str = coat_dt.strftime('%y%m%d')

                    # 製品IDを逆引き（メモリ上で実行）
                    product_key = (base, adp, lr, color)
                    prd_id = product_master.get(product_key)
                    
                    if not prd_id:
                        error_msg = f'シート「{sheet_name}」行{row}: 製品IDが見つかりません: ベース={base}, 加入度数={adp}, L/R={lr}, 色={color}'
                        error_messages.append(error_msg)
                        error_count += 1
                        continue

                    # 既存データチェック（メモリ上で実行）
                    if (prd_id, lot) in existing_lots:
                        error_msg = f'シート「{sheet_name}」行{row}: 一意制約違反: 製品ID={prd_id}, ロット={lot}のデータが既に存在します'
                        error_messages.append(error_msg)
                        error_count += 1
                        continue

                    data_to_insert.append({
                        'prd_id': prd_id,
                        'lot': lot,  # 文字列として保存
                        'qty': quantity,
                        'coat_date': coat_dt_str,
                        'proc_hardcoat': DatabaseConstants.PROC_HARD_COAT,
                        'flg_not_shipped': DatabaseConstants.BPDD_FLG_NOT_SHIPPED,
                        'bshk_id': bshk_id
                    })
                    
                except (ValueError, TypeError) as e:
                    error_msg = f'シート「{sheet_name}」行{row}: データ形式エラー: {str(e)}'
                    error_messages.append(error_msg)
                    error_count += 1
                    continue
                except Exception as e:
                    error_msg = f'シート「{sheet_name}」行{row}: 予期しないエラー: {str(e)}'
                    error_messages.append(error_msg)
                    error_count += 1
                    continue

            # データを一括挿入
            if data_to_insert:
                try:
                    sql = text("""
                        INSERT INTO BPRD_DAT
                        (BPDD_PRD_ID, BPDD_LOT, BPDD_QTY, BPDD_CRT, BPDD_PROC, BPDD_FLG)
                        VALUES
                        (:prd_id, :lot, :qty, :coat_date, :proc_hardcoat, :flg_not_shipped)
                    """)
                    session.execute(sql, data_to_insert)

                    for data in data_to_insert:
                        if data['bshk_id'] == 0:
                            continue
                    
                        # SQL文を動的に生成
                        sql_text = """
                            UPDATE BSHK_DAT
                            SET BSHK_FLG = :bshk_flg_processed
                            WHERE BSHK_ID = :bshk_id
                            AND BSHK_FLG = :bshk_flg_not_shipped
                        """
                        session.execute(text(sql_text), {'bshk_id': data['bshk_id'], 'bshk_flg_processed': DatabaseConstants.BSHK_FLG_PROCESSED, 'bshk_flg_not_shipped': DatabaseConstants.BSHK_FLG_NOT_SHIPPED})
                    
                    processed_count += len(data_to_insert)
                    
                except Exception as e:
                    error_msg = f'シート「{sheet_name}」: データベース登録エラー: {str(e)}'
                    error_messages.append(error_msg)
                    error_count += 1
                    session.rollback()
                    continue

        # 最終コミット
        if processed_count > 0:
            session.commit()
            
        # 結果メッセージの構築
        if processed_count > 0 and error_count == 0:
            flash(f'ファイルのインポートが完了しました。処理件数: {processed_count}件', 'success')
        elif processed_count > 0 and error_count > 0:
            flash(f'部分的なインポートが完了しました。成功: {processed_count}件, エラー: {error_count}件', 'warning')
        else:
            flash(f'インポートに失敗しました。エラー件数: {error_count}件', 'error')
            
        # エラーメッセージがある場合は詳細を表示
        if error_messages:
            error_details = '\n'.join(error_messages[:10])  # 最初の10件のみ表示
            if len(error_messages) > 10:
                error_details += f'\n... 他 {len(error_messages) - 10}件のエラーがあります'
            flash(f'エラー詳細:\n{error_details}', 'error')
            
        return redirect(url_for('hardcoat_read_excel'))

    except Exception as e:
        session.rollback()
        tb = traceback.format_exc()
        flash(f'Excelファイルのインポート中にエラーが発生しました: {str(e)}', 'error')
        flash(f'エラー詳細:\n{tb}', 'error')
        return redirect(url_for('hardcoat_read_excel'))
    finally:
        session.close()
        # 一時ファイルを削除（元のまま）
        try:
            if os.path.exists(temp_path):
                os.remove(temp_path)
        except Exception as e:
            print(f"一時ファイルの削除に失敗: {str(e)}")

